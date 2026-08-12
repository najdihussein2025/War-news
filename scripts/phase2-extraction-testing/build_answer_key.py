import json
import re
import zipfile
from collections import Counter
from pathlib import Path
from xml.etree import ElementTree as ET


SCRIPT_DIR = Path(__file__).resolve().parent
PROJECT_ROOT = SCRIPT_DIR.parents[1]
WORKBOOK_PATH = PROJECT_ROOT / "Data" / "Database Sample.xlsx"
OUTPUT_PATH = SCRIPT_DIR / "answer_key.json"
SAMPLE_TEXTS_DIR = SCRIPT_DIR / "sample_texts"

NS_MAIN = {"x": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
NS_REL = {"r": "http://schemas.openxmlformats.org/package/2006/relationships"}


CATEGORY_GROUPS = {
    "army": {
        "label": "Lebanese Army",
        "flags": ["LA"],
        "fields": ["LA", "LA_DID", "LA_Bldg", "LA_V", "LAM_D", "LAM_I", "LAF_D", "LAF_I", "LA_TD", "LA_TI"],
    },
    "unifil": {
        "label": "UNIFIL",
        "flags": ["UNIFIL"],
        "fields": ["UNIFIL", "UN_DID", "UN_Bldg", "UN_V", "UNM_D", "UNM_I", "UNF_D", "UNF_I", "UN_TD", "UN_TI"],
    },
    "municipality": {
        "label": "Municipality",
        "flags": ["MUNI"],
        "fields": ["MUNI", "MUNI_DID", "MUNI_Bldg", "MUNI_Empl", "MUNIM_D", "MUNIM_I", "MUNIF_D", "MUNIF_I", "MUNI_TD", "MUNI_TI"],
    },
    "school_university": {
        "label": "School/University",
        "flags": ["School", "Uni"],
        "fields": ["School", "Sch_DID", "School_Name", "School_Damage_Level", "Uni", "Uni_DID", "Uni_Name"],
    },
    "religious_cultural": {
        "label": "Religious & cultural",
        "flags": ["Church", "Mosque", "Ceme", "Releg", "Archeo"],
        "fields": ["Church", "Chu_DID", "Chu_N", "Mosque", "Mos_DID", "Mosque_N", "Ceme", "Ceme_DID", "Ceme_N", "Releg", "Releg_DID", "Releg_N", "Archeo", "Arch_DID", "Arch_N"],
    },
    "hospital": {
        "label": "Hospital",
        "flags": ["Hosp"],
        "fields": ["Hosp", "Hos_DID", "Hos_Status", "Hos_N", "Hos_Damage_Level", "Nbr_Evap", "HosM_D", "HosM_I", "HosF_D", "HosF_I", "HosD", "HosI"],
    },
    "health_center": {
        "label": "Health Center",
        "flags": ["HC"],
        "fields": ["HC", "HC_Rela", "HC_DID", "HC_Damage_Level", "HCM_D", "HCM_I", "HCF_D", "HCF_I", "HCD", "HCI"],
    },
    "emergency_civil_defense": {
        "label": "Emergency/Civil Defense",
        "flags": ["Emer", "E_Cars"],
        "fields": ["Emer", "E_Cars", "Car_Nbr", "Emer_Rela", "Emer_D", "Emer_I"],
    },
    "press": {
        "label": "Press",
        "flags": ["Press"],
        "fields": ["Press", "Channel", "Press_DID", "PressM_D", "PressM_I", "PressF_D", "PressF_I", "PressD", "PressI"],
    },
    "government_building": {
        "label": "Government building",
        "flags": ["Gov", "Gov_Bui"],
        "fields": ["Gov", "Gov_Bui", "Gov_N", "GB_DID", "GBM_D", "GBM_I", "GBF_D", "GBF_I", "GBD", "GBI"],
    },
    "road_bridge": {
        "label": "Road/Bridge",
        "flags": ["Road", "Bridge"],
        "fields": ["Road", "Road_D_ID", "Road_Blocked", "Road_Name", "Bridge", "Bridge_Blocked", "Bridge_Name"],
    },
    "vehicles": {
        "label": "Vehicles",
        "flags": ["Car", "Moto", "Con_Veh", "Excavator", "Bulldozer", "Camion", "Bobcat", "Tracteur"],
        "fields": ["Car", "CarD", "CarI", "CarM_D", "CarM_I", "CarF_D", "CarF_I", "CarC_D", "CarC_I", "Moto", "Moto_DID", "Moto_D", "Moto_I", "Con_Veh", "Con_D", "Con_I", "Excavator", "Bulldozer", "Camion", "Bobcat", "Tracteur", "Total_Con"],
    },
    "crossings_other": {
        "label": "Crossings & other",
        "flags": ["Crossing", "Litani", "Zahrani", "Drone_F", "Water", "Electric", "Olives_Trees_D", "Mjnoub", "Other"],
        "fields": ["Crossing", "Litani", "Zahrani", "Drone_F", "Water", "Water_DID", "Water_Type", "Electric", "Electric_DID", "Electric_Type", "Olives_Trees_D", "Mjnoub", "MJ_DID", "Other", "Other_DID", "Other_Type", "Other_D", "Other_I"],
    },
    "warning_classification": {
        "label": "Warning & classification",
        "flags": ["No_Warning", "Warning", "Genocide", "Building", "Apart"],
        "fields": ["No_Warning", "Warning", "Genocide", "Building", "Apart"],
    },
}

CASUALTY_FIELDS = ["Total_D", "Total_Inj", "Death", "Injuries", "Male_D", "Male_I", "female_D", "female_I", "Children_D", "Children_I"]
TOP_LEVEL_FIELDS = ["ACS_Name", "Action_E", "Action_A"]
FIELD_ALIASES = {
    "School_Damage_Level": ["Damage_Level"],
    "Hos_Damage_Level": ["Dam_Level"],
    "HC_Damage_Level": ["Dam_Level__2"],
    "Nbr_Evap": ["NBR_EVAP"],
    "Car_Nbr": ["Car_nbr"],
    "Emer_D": ["EmerD"],
    "Emer_I": ["EmerI"],
    "Road_Blocked": ["Blocked"],
    "Bridge_Blocked": ["Blocked__2"],
    "Total_Con": ["Total _Con"],
    "Olives_Trees_D": ["Olives Trees_D"],
    "Mjnoub": ["MJnoub"],
    "Other_D": ["OtherD"],
}


def normalize_header(value):
    return re.sub(r"\s+", " ", str(value or "")).strip()


def normalize_key(value):
    return normalize_header(value).lower().replace(" ", "_")


def make_unique_headers(headers):
    counts = Counter()
    unique = []
    for index, header in enumerate(headers, start=1):
        base = normalize_header(header) or f"Column_{index}"
        key = normalize_key(base)
        counts[key] += 1
        if counts[key] == 1:
            unique.append(base)
        else:
            unique.append(f"{base}__{counts[key]}")
    return unique


def is_present(value):
    if value is None:
        return False
    if isinstance(value, str):
        stripped = value.strip()
        if not stripped:
            return False
        return stripped.lower() not in {"0", "0.0", "nan", "none", "null", "no", "n/a"}
    if isinstance(value, (int, float)):
        return value != 0
    return True


def clean_value(value):
    if isinstance(value, float) and value.is_integer():
        return int(value)
    if isinstance(value, str):
        return value.strip()
    return value


def column_index(cell_ref):
    letters = re.sub(r"[^A-Z]", "", cell_ref.upper())
    index = 0
    for letter in letters:
        index = index * 26 + ord(letter) - ord("A") + 1
    return index - 1


def read_shared_strings(zf):
    path = "xl/sharedStrings.xml"
    if path not in zf.namelist():
        return []
    root = ET.fromstring(zf.read(path))
    values = []
    for si in root.findall("x:si", NS_MAIN):
        parts = [node.text or "" for node in si.findall(".//x:t", NS_MAIN)]
        values.append("".join(parts))
    return values


def find_sheet_path(zf, sheet_name):
    workbook = ET.fromstring(zf.read("xl/workbook.xml"))
    rels = ET.fromstring(zf.read("xl/_rels/workbook.xml.rels"))
    rel_by_id = {rel.attrib["Id"]: rel.attrib["Target"] for rel in rels.findall("r:Relationship", NS_REL)}
    selected = None
    for sheet in workbook.findall(".//x:sheet", NS_MAIN):
        if sheet.attrib.get("name") == sheet_name:
            selected = sheet
            break
    if selected is None:
        selected = workbook.find(".//x:sheet", NS_MAIN)
    if selected is None:
        raise ValueError("Workbook has no sheets")
    rel_id = selected.attrib["{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id"]
    target = rel_by_id[rel_id].lstrip("/")
    if not target.startswith("xl/"):
        target = "xl/" + target
    return target, selected.attrib.get("name", sheet_name)


def parse_cell(cell, shared_strings):
    cell_type = cell.attrib.get("t")
    value_node = cell.find("x:v", NS_MAIN)
    if cell_type == "inlineStr":
        return "".join(node.text or "" for node in cell.findall(".//x:t", NS_MAIN))
    if value_node is None:
        return None
    raw = value_node.text
    if raw is None:
        return None
    if cell_type == "s":
        return shared_strings[int(raw)]
    if cell_type == "b":
        return int(raw)
    try:
        number = float(raw)
        return int(number) if number.is_integer() else number
    except ValueError:
        return raw


def read_xlsx_rows(path, sheet_name="Sheet1"):
    with zipfile.ZipFile(path) as zf:
        shared_strings = read_shared_strings(zf)
        sheet_path, actual_sheet = find_sheet_path(zf, sheet_name)
        root = ET.fromstring(zf.read(sheet_path))
        rows = []
        for row in root.findall(".//x:sheetData/x:row", NS_MAIN):
            values = []
            for cell in row.findall("x:c", NS_MAIN):
                ref = cell.attrib.get("r", "")
                idx = column_index(ref)
                while len(values) <= idx:
                    values.append(None)
                values[idx] = parse_cell(cell, shared_strings)
            rows.append(values)
        return rows, actual_sheet


def load_rows():
    try:
        import openpyxl

        workbook = openpyxl.load_workbook(WORKBOOK_PATH, read_only=True, data_only=True)
        sheet = workbook["Sheet1"] if "Sheet1" in workbook.sheetnames else workbook[workbook.sheetnames[0]]
        return [[cell.value for cell in row] for row in sheet.iter_rows()], sheet.title
    except ImportError:
        return read_xlsx_rows(WORKBOOK_PATH)


def row_to_record(headers, row):
    return {headers[i]: clean_value(row[i]) if i < len(row) else None for i in range(len(headers))}


def first_existing(record, names):
    by_norm = {normalize_key(key): key for key in record}
    for name in names:
        key = by_norm.get(normalize_key(name))
        if key is not None:
            return key
    return None


def collect_fields(record, fields):
    by_norm = {normalize_key(key): key for key in record}
    output = {}
    for field in fields:
        candidates = [field] + FIELD_ALIASES.get(field, [])
        key = None
        for candidate in candidates:
            key = by_norm.get(normalize_key(candidate))
            if key is not None:
                break
        if key is not None and is_present(record[key]):
            output[field] = clean_value(record[key])
    return output


def group_is_present(record, group):
    by_norm = {normalize_key(key): key for key in record}
    for flag in group["flags"]:
        key = by_norm.get(normalize_key(flag))
        if key is not None and is_present(record[key]):
            return True
    return bool(collect_fields(record, group["fields"]))


def main():
    rows, sheet_name = load_rows()
    if not rows:
        raise RuntimeError(f"No rows found in {WORKBOOK_PATH}")

    headers = make_unique_headers(rows[0])
    khabar_col = first_existing({header: None for header in headers}, ["Khabar"])
    if not khabar_col:
        raise RuntimeError("Could not find Khabar column")

    SAMPLE_TEXTS_DIR.mkdir(parents=True, exist_ok=True)
    for old_sample in SAMPLE_TEXTS_DIR.glob("sample_*.txt"):
        old_sample.unlink()

    answer_key = []
    category_distribution = Counter()
    processed = max(len(rows) - 1, 0)

    for excel_row_number, raw_row in enumerate(rows[1:], start=2):
        record = row_to_record(headers, raw_row)
        khabar_text = clean_value(record.get(khabar_col))
        if not is_present(khabar_text):
            continue

        sample_id = excel_row_number
        expected = {}

        for field in TOP_LEVEL_FIELDS:
            values = collect_fields(record, [field])
            expected.update(values)

        casualties = collect_fields(record, CASUALTY_FIELDS)
        if casualties:
            expected["casualties"] = casualties

        categories = {}
        for category_key, group in CATEGORY_GROUPS.items():
            if group_is_present(record, group):
                fields = collect_fields(record, group["fields"])
                categories[category_key] = {
                    "label": group["label"],
                    "fields": fields,
                }
                category_distribution[group["label"]] += 1
        if categories:
            expected["categories"] = categories

        answer_key.append(
            {
                "sample_id": sample_id,
                "source_row": excel_row_number,
                "khabar_text": str(khabar_text),
                "expected": expected,
            }
        )
        (SAMPLE_TEXTS_DIR / f"sample_{sample_id}.txt").write_text(str(khabar_text), encoding="utf-8")

    OUTPUT_PATH.write_text(json.dumps(answer_key, ensure_ascii=False, indent=2), encoding="utf-8")

    print(f"Workbook: {WORKBOOK_PATH}")
    print(f"Sheet: {sheet_name}")
    print(f"Total rows processed: {processed}")
    print(f"Rows with non-empty Khabar: {len(answer_key)}")
    print("Category distribution:")
    for label, count in category_distribution.most_common():
        print(f"- {label}: {count} rows")
    print(f"Answer key written to: {OUTPUT_PATH}")
    print(f"Sample texts written to: {SAMPLE_TEXTS_DIR}")


if __name__ == "__main__":
    main()
