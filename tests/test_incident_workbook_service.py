from __future__ import annotations

from datetime import date, time
from io import BytesIO
from pathlib import Path
from uuid import uuid4

from openpyxl import Workbook, load_workbook

from app.news.models import Condition, Incident, IncidentDetail, Village
from app.news.services.incident_workbook_service import IncidentWorkbookService
from app.sources.models import Source, SourceType


class _ScalarResult:
    def __init__(self, items):
        self._items = list(items)

    def all(self):
        return list(self._items)


class _FakeSession:
    def __init__(self, *, villages=None, conditions=None, sources=None):
        self._tables = {
            Village: villages or [],
            Condition: conditions or [],
            Source: sources or [],
        }
        self.added: list[object] = []
        self.commits = 0
        self.rollbacks = 0

    def scalars(self, query):
        entity = query.column_descriptions[0]["entity"]
        return _ScalarResult(self._tables.get(entity, []))

    def add(self, obj):
        if isinstance(obj, Incident) and obj.id is None:
            obj.id = uuid4()
        self.added.append(obj)

    def flush(self):
        return None

    def commit(self):
        self.commits += 1

    def rollback(self):
        self.rollbacks += 1


class _SchemaAwareSession(_FakeSession):
    def __init__(self, *, bind, villages=None, conditions=None, sources=None):
        super().__init__(villages=villages, conditions=conditions, sources=sources)
        self._bind = bind

    def get_bind(self):
        return self._bind


def _legacy_headers() -> list[object]:
    workbook_path = Path("Data") / "Database Sample.xlsx"
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    sheet = workbook.active
    return list(next(sheet.iter_rows(min_row=1, max_row=1, values_only=True)))


def _set_header_value(row: list[object], headers: list[object], header: str, value: object, occurrence: int = 1) -> None:
    seen = 0
    for index, current in enumerate(headers):
        current_text = "" if current is None else str(current)
        if current_text == header:
            seen += 1
            if seen == occurrence:
                row[index] = value
                return
    raise AssertionError(f"Header {header!r} occurrence {occurrence} not found.")


def _build_workbook(row: list[object]) -> BytesIO:
    workbook = Workbook()
    sheet = workbook.active
    headers = _legacy_headers()
    sheet.append(headers)
    sheet.append(row)
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


def test_import_workbook_maps_legacy_row_fields() -> None:
    headers = _legacy_headers()
    row = [None] * len(headers)
    row[0] = "Lookup-only village text"
    _set_header_value(row, headers, "ACS_Code", 12345)
    _set_header_value(row, headers, "Month", "August")
    _set_header_value(row, headers, "Action_E", "Shelling")
    _set_header_value(row, headers, "Action_A", "قصف")
    _set_header_value(row, headers, "Khabar", "Workbook incident")
    _set_header_value(row, headers, "Source", "Daily Source")
    _set_header_value(row, headers, "Time", time(8, 30))
    _set_header_value(row, headers, "Date", date(2026, 8, 21))
    _set_header_value(row, headers, "Total_D", 3)
    _set_header_value(row, headers, "Total_Inj", 4)
    _set_header_value(row, headers, "Death", 1)
    _set_header_value(row, headers, "Injuries", 2, occurrence=1)
    _set_header_value(row, headers, "Male_D", 1)
    _set_header_value(row, headers, "LA", 1)
    _set_header_value(row, headers, "LA_DID", "D")
    _set_header_value(row, headers, "LA_TD", 1)
    _set_header_value(row, headers, "NOTE", "Primary note")
    _set_header_value(row, headers, "MOH", "MOH value")
    _set_header_value(row, headers, "Worker Name", "Worker")
    _set_header_value(row, headers, "Links", "https://source-1")
    _set_header_value(row, headers, "Links", "https://source-2", occurrence=2)
    _set_header_value(row, headers, "Martyrs", "1 martyr")
    _set_header_value(row, headers, "Injuries", 7, occurrence=2)
    _set_header_value(row, headers, "Note", "Secondary note", occurrence=1)
    _set_header_value(row, headers, "Note", "Tertiary note", occurrence=2)

    session = _FakeSession(
        villages=[Village(id=11, acs_code=12345, ref_name_en="Village")],
        conditions=[Condition(id=22, action_en="Shelling", action_ar="قصف")],
        sources=[Source(id=33, type=SourceType.api, name="Daily Source", config={})],
    )

    summary = IncidentWorkbookService(session).import_workbook(
        _build_workbook(row),
        created_by=uuid4(),
    )

    incidents = [item for item in session.added if isinstance(item, Incident)]
    details = [item for item in session.added if isinstance(item, IncidentDetail)]

    assert summary.processed == 1
    assert summary.succeeded == 1
    assert summary.failed == 0
    assert summary.row_errors == []
    assert session.commits == 1
    assert len(incidents) == 1
    assert len(details) == 1

    incident = incidents[0]
    detail = details[0]
    assert incident.village_id == 11
    assert incident.condition_id == 22
    assert incident.source_id == 33
    assert incident.event_month == "August"
    assert incident.khabar == "Workbook incident"
    assert incident.total_deaths == 3
    assert incident.total_injuries == 4
    assert incident.deaths == 1
    assert incident.injuries == 2
    assert incident.injuries_extra == 7
    assert incident.note == "Primary note"
    assert incident.note_extra == "Secondary note"
    assert incident.note_extra_2 == "Tertiary note"
    assert incident.source_link == "https://source-1"
    assert incident.source_link_2 == "https://source-2"
    assert incident.moh == "MOH value"
    assert incident.worker_name == "Worker"
    assert incident.martyrs == "1 martyr"
    assert detail.male_d == 1
    assert detail.la is True
    assert detail.la_did == "D"
    assert detail.la_td == 1


def test_import_workbook_leaves_lookup_fields_null_when_unmatched() -> None:
    headers = _legacy_headers()
    row = [None] * len(headers)
    row[0] = "Lookup-only village text"
    _set_header_value(row, headers, "ACS_Code", 99999)
    _set_header_value(row, headers, "Month", "August")
    _set_header_value(row, headers, "Action_E", "Unknown Action")
    _set_header_value(row, headers, "Source", "Unknown Source")
    _set_header_value(row, headers, "Khabar", "Workbook incident")
    _set_header_value(row, headers, "Time", time(8, 30))
    _set_header_value(row, headers, "Date", date(2026, 8, 21))

    session = _FakeSession()

    summary = IncidentWorkbookService(session).import_workbook(_build_workbook(row))

    incident = next(item for item in session.added if isinstance(item, Incident))
    assert summary.processed == 1
    assert summary.succeeded == 1
    assert summary.failed == 0
    assert incident.village_id is None
    assert incident.condition_id is None
    assert incident.source_id is None


def test_import_workbook_allows_missing_optional_legacy_columns() -> None:
    optional_headers = {"Injuries__2", "Links__2", "Martyrs", "Note", "Note__2"}
    headers = IncidentWorkbookService._make_unique_headers(_legacy_headers())
    headers = [header for header in headers if header not in optional_headers]
    row = [None] * len(headers)
    row[headers.index("Khabar")] = "Workbook incident"
    row[headers.index("Date")] = date(2026, 8, 21)

    workbook = Workbook()
    sheet = workbook.active
    sheet.append(headers)
    sheet.append(row)
    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    session = _FakeSession()
    summary = IncidentWorkbookService(session).import_workbook(output)

    incident = next(item for item in session.added if isinstance(item, Incident))
    assert summary.succeeded == 1
    assert summary.failed == 0
    assert incident.injuries_extra is None
    assert incident.source_link_2 is None
    assert incident.martyrs is None
    assert incident.note_extra is None
    assert incident.note_extra_2 is None


def test_import_workbook_requires_note_extra_2_schema_column(monkeypatch) -> None:
    headers = _legacy_headers()
    row = [None] * len(headers)
    _set_header_value(row, headers, "Khabar", "Workbook incident")
    _set_header_value(row, headers, "Date", date(2026, 8, 21))

    class _Inspector:
        @staticmethod
        def get_columns(_table_name):
            return [{"name": "id"}, {"name": "note"}, {"name": "note_extra"}]

    monkeypatch.setattr(
        "app.news.services.incident_workbook_service.inspect",
        lambda _bind: _Inspector(),
    )

    session = _SchemaAwareSession(bind=object())

    try:
        IncidentWorkbookService(session).import_workbook(_build_workbook(row))
        raise AssertionError("Expected import to fail when note_extra_2 is missing.")
    except ValueError as exc:
        assert str(exc) == (
            "Database schema is out of date for incident imports. "
            "Missing incidents columns: note_extra_2. "
            "Apply the latest Alembic migrations before importing incidents."
        )
