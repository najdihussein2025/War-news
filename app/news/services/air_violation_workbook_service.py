from datetime import date, datetime, time
from io import BytesIO
from typing import BinaryIO

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.accounts import models as account_models  # noqa: F401
from app.logs import models as log_models  # noqa: F401
from app.news.models import AirViolation, Condition
from app.sources.models import Source, SourceType

HEADERS = ["Caza", "Month", "Action_E", "Action_A", "Khabar", "Source", "Time", "Date", "Note 1", "Note 2", "Link"]
AIR_CONDITION_IDS = {35, 36, 38}


class AirViolationWorkbookService:
    def __init__(self, db: Session) -> None:
        self.db = db

    def import_workbook(self, stream: BinaryIO) -> dict[str, int]:
        workbook = load_workbook(stream, read_only=True, data_only=True)
        sheet = workbook.active
        headers = [str(value).strip() if value is not None else "" for value in next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))]
        missing = [header for header in HEADERS if header not in headers]
        if missing:
            raise ValueError(f"Missing required columns: {', '.join(missing)}")
        indexes = {header: headers.index(header) for header in HEADERS}
        conditions = list(self.db.scalars(select(Condition).where(Condition.id.in_(AIR_CONDITION_IDS))).all())
        by_action = {condition.action_en.strip().casefold(): condition for condition in conditions}
        by_action.update({condition.action_ar.strip().casefold(): condition for condition in conditions})
        imported = skipped = failed = 0
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not any(value not in (None, "") for value in row):
                continue
            try:
                action_en = str(row[indexes["Action_E"]] or "").strip()
                action_ar = str(row[indexes["Action_A"]] or "").strip()
                condition = by_action.get(action_en.casefold()) or by_action.get(action_ar.casefold())
                if condition is None:
                    failed += 1
                    continue
                source_name = str(row[indexes["Source"]] or "Manual import").strip()
                source = self.db.scalar(select(Source).where(Source.name == source_name))
                if source is None:
                    source = Source(type=SourceType.manual, name=source_name, config={}, is_active=True)
                    self.db.add(source)
                    self.db.flush()
                event_date = self._date(row[indexes["Date"]])
                event_time = self._time(row[indexes["Time"]])
                khabar = str(row[indexes["Khabar"]] or "").strip()
                caza = str(row[indexes["Caza"]] or "").strip() or None
                duplicate = self.db.scalar(select(AirViolation.id).where(
                    AirViolation.condition_id == condition.id,
                    AirViolation.source_id == source.id,
                    AirViolation.event_date == event_date,
                    AirViolation.event_time == event_time,
                    AirViolation.khabar == khabar,
                ))
                if duplicate is not None:
                    skipped += 1
                    continue
                self.db.add(AirViolation(
                    condition_id=condition.id,
                    source_id=source.id,
                    caza_en=caza,
                    event_month=str(row[indexes["Month"]] or event_date.strftime("%B")),
                    event_date=event_date,
                    event_time=event_time,
                    khabar=khabar,
                    note_1=self._optional(row[indexes["Note 1"]]),
                    note_2=self._optional(row[indexes["Note 2"]]),
                    source_link=self._optional(row[indexes["Link"]]),
                ))
                imported += 1
            except (TypeError, ValueError):
                failed += 1
        self.db.commit()
        return {"imported": imported, "skipped": skipped, "failed": failed}

    def export_workbook(self) -> BytesIO:
        rows = self.db.execute(
            select(AirViolation, Condition, Source)
            .join(Condition, Condition.id == AirViolation.condition_id)
            .join(Source, Source.id == AirViolation.source_id)
            .order_by(AirViolation.event_date.desc(), AirViolation.event_time.desc().nullslast())
        ).all()
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Air Violations"
        sheet.append(HEADERS)
        for cell in sheet[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="FFE699" if cell.column <= 8 or cell.column == 11 else "BDD7EE")
            cell.alignment = Alignment(horizontal="center")
        for violation, condition, source in rows:
            sheet.append([
                violation.caza_en or violation.caza_ar,
                violation.event_month,
                condition.action_en,
                condition.action_ar,
                violation.khabar,
                source.name,
                violation.event_time,
                violation.event_date,
                violation.note_1,
                violation.note_2,
                violation.source_link,
            ])
        widths = [20, 14, 28, 28, 60, 28, 14, 14, 35, 35, 45]
        for index, width in enumerate(widths, start=1):
            sheet.column_dimensions[chr(64 + index)].width = width
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        output = BytesIO()
        workbook.save(output)
        output.seek(0)
        return output

    @staticmethod
    def _optional(value: object) -> str | None:
        return str(value).strip() if value not in (None, "") else None

    @staticmethod
    def _date(value: object) -> date:
        if isinstance(value, datetime):
            return value.date()
        if isinstance(value, date):
            return value
        return datetime.fromisoformat(str(value).strip()).date()

    @staticmethod
    def _time(value: object) -> time | None:
        if value in (None, ""):
            return None
        if isinstance(value, datetime):
            return value.time().replace(tzinfo=None)
        if isinstance(value, time):
            return value.replace(tzinfo=None)
        return time.fromisoformat(str(value).strip())
