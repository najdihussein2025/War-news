from __future__ import annotations

from datetime import date, datetime, time
from typing import Any, BinaryIO
from uuid import UUID

from openpyxl import load_workbook
from sqlalchemy import inspect
from sqlalchemy import select
from sqlalchemy.orm import Session
from sqlalchemy.sql.sqltypes import Boolean, Enum, Integer

from app.news.dtos import WorkbookImportRowErrorDTO, WorkbookImportSummaryDTO
from app.news.models import Condition, Incident, IncidentDetail, Village
from app.sources.models import Source

LOOKUP_ONLY_HEADERS: tuple[str, ...] = (
    "Column_1",
    "ACS_Code",
    "ACS_Name",
    "C_Ref_Eng",
    "C_Ref_Arab",
    "Gove_E",
    "Gove_A",
    "Dist_E",
    "Dist_A",
    "Action_E",
    "Action_A",
    "Source",
)

INCIDENT_FIELD_MAP: dict[str, str] = {
    "Month": "event_month",
    "Khabar": "khabar",
    "MOH": "moh",
    "Worker Name": "worker_name",
    "Martyrs": "martyrs",
    "Links": "source_link",
    "Links__2": "source_link_2",
    "NOTE": "note",
    "Note": "note_extra",
    "Note__2": "note_extra_2",
}

INCIDENT_INT_FIELD_MAP: dict[str, str] = {
    "Total_D": "total_deaths",
    "Total_Inj": "total_injuries",
    "Death": "deaths",
    "Injuries": "injuries",
    "Injuries__2": "injuries_extra",
}

INCIDENT_DETAIL_FIELD_MAP: dict[str, str] = {
    "Male_D": "male_d",
    "Male_I": "male_i",
    "female_D": "female_d",
    "female_I": "female_i",
    "Children_D": "children_d",
    "Children_I": "children_i",
    "Obs_Duties": "obs_duties",
    "ISF_GS": "isf_gs",
    "Fire": "fire",
    "Arrested": "arrested",
    "Lib_Y_N": "lib_y_n",
    "LA": "la",
    "LA_DID": "la_did",
    "LA_Bldg": "la_bldg",
    "LA_V": "la_v",
    "LAM_D": "lam_d",
    "LAM_I": "lam_i",
    "LAF_D": "laf_d",
    "LAF_I": "laf_i",
    "LA_TD": "la_td",
    "LA_TI": "la_ti",
    "Unifil": "unifil",
    "Un_DID": "un_did",
    "Un_Bldg": "un_bldg",
    "Un_V": "un_v",
    "UNM_D": "unm_d",
    "UNM_I": "unm_i",
    "UNF_D": "unf_d",
    "UNF_I": "unf_i",
    "UN_TD": "un_td",
    "UN_TI": "un_ti",
    "Muni": "muni",
    "Muni_DID": "muni_did",
    "Muni_bldg": "muni_bldg",
    "Muni_empl": "muni_empl",
    "MuniM_D": "munim_d",
    "MuniM_I": "munim_i",
    "MuniF_D": "munif_d",
    "MuniF_I": "munif_i",
    "Muni_TD": "muni_td",
    "Muni_TI": "muni_ti",
    "School": "school",
    "Sch_DID": "sch_did",
    "School_Name": "school_name",
    "Damage_Level": "school_damage_level",
    "Uni": "uni",
    "Uni_DID": "uni_did",
    "UNI_Name": "uni_name",
    "Church": "church",
    "Chu_DID": "chu_did",
    "Chu_N": "chu_n",
    "Mosque": "mosque",
    "Mos_DID": "mos_did",
    "Mosque_N": "mosque_n",
    "Ceme": "ceme",
    "Ceme_DID": "ceme_did",
    "Ceme_N": "ceme_n",
    "Releg": "releg",
    "Releg_DID": "releg_did",
    "Releg_N": "releg_n",
    "Archeo": "archeo",
    "Arch_DID": "arch_did",
    "Arch_N": "arch_n",
    "Hosp": "hosp",
    "Hos_DID": "hos_did",
    "Hos_Status": "hos_status",
    "Hos_N": "hos_n",
    "Dam_Level": "hos_damage_level",
    "NBR_EVAP": "nbr_evap",
    "HosM_D": "hosm_d",
    "HosM_I": "hosm_i",
    "HosF_D": "hosf_d",
    "HosF_I": "hosf_i",
    "HosD": "hosd",
    "HosI": "hosi",
    "HC": "hc",
    "HC_Rela": "hc_rela",
    "HC_DID": "hc_did",
    "Dam_Level__2": "hc_damage_level",
    "HCM_D": "hcm_d",
    "HCM_I": "hcm_i",
    "HCF_D": "hcf_d",
    "HCF_I": "hcf_i",
    "HCD": "hcd",
    "HCI": "hci",
    "Emer": "emer",
    "E_Cars": "e_cars",
    "Car_nbr": "car_nbr",
    "Emer_Rela": "emer_rela",
    "EmerD": "emer_d",
    "EmerI": "emer_i",
    "Press": "press",
    "Channel": "channel",
    "Press_DID": "press_did",
    "PressM_D": "pressm_d",
    "PressM_I": "pressm_i",
    "PressF_D": "pressf_d",
    "PressF_I": "pressf_i",
    "PressD": "pressd",
    "PressI": "pressi",
    "Gov": "gov",
    "Gov_Bui": "gov_bui",
    "Gov_N": "gov_n",
    "GB_DID": "gb_did",
    "GBM_D": "gbm_d",
    "GBM_I": "gbm_i",
    "GBF_D": "gbf_d",
    "GBF_I": "gbf_i",
    "GBD": "gbd",
    "GBI": "gbi",
    "Road": "road",
    "Road_D_ID": "road_d_id",
    "Blocked": "road_blocked",
    "Road_Name": "road_name",
    "Bridge": "bridge",
    "Blocked__2": "bridge_blocked",
    "Bridge_Name": "bridge_name",
    "Car": "car",
    "CarD": "card",
    "CarI": "cari",
    "CarM_D": "carm_d",
    "CarM_I": "carm_i",
    "CarF_D": "carf_d",
    "CarF_I": "carf_i",
    "CarC_D": "carc_d",
    "CarC_I": "carc_i",
    "Moto": "moto",
    "Moto_DID": "moto_did",
    "Moto_D": "moto_d",
    "Moto_I": "moto_i",
    "Con_Veh": "con_veh",
    "Con_D": "con_d",
    "Con_I": "con_i",
    "Excavator": "excavator",
    "Bulldozer": "bulldozer",
    "Camion": "camion",
    "Bobcat": "bobcat",
    "Total _Con": "total_con",
    "Tracteur": "tracteur",
    "Crossing": "crossing",
    "Litani": "litani",
    "Zahrani": "zahrani",
    "Drone_F": "drone_f",
    "Water": "water",
    "Water_DID": "water_did",
    "Water_Type": "water_type",
    "Electric": "electric",
    "Electric_DID": "electric_did",
    "Electric_Type": "electric_type",
    "Olives Trees_D": "olives_trees_d",
    "MJnoub": "mjnoub",
    "MJ_DID": "mj_did",
    "Other": "other",
    "Other_DID": "other_did",
    "Other_type": "other_type",
    "OtherD": "other_d",
    "Other_I": "other_i",
    "No_warning": "no_warning",
    "Warning": "warning",
    "Genocide": "genocide",
    "Building": "building",
    "Apart": "apart",
}

OPTIONAL_HEADERS: frozenset[str] = frozenset(
    {
        "Injuries__2",
        "Links__2",
        "Martyrs",
        "Note",
        "Note__2",
    }
)

REQUIRED_HEADERS: frozenset[str] = frozenset(
    set(LOOKUP_ONLY_HEADERS)
    | set(INCIDENT_FIELD_MAP)
    | set(INCIDENT_INT_FIELD_MAP)
    | {"Time", "Date"}
    | set(INCIDENT_DETAIL_FIELD_MAP)
) - OPTIONAL_HEADERS


class IncidentWorkbookService:
    def __init__(self, db: Session) -> None:
        self.db = db

    def import_workbook(
        self,
        stream: BinaryIO,
        *,
        created_by: UUID | None = None,
    ) -> WorkbookImportSummaryDTO:
        self._ensure_schema_compatible()
        workbook = load_workbook(stream, read_only=True, data_only=True)
        sheet = workbook.active
        rows = sheet.iter_rows(values_only=True)
        try:
            raw_headers = next(rows)
        except StopIteration as exc:
            raise ValueError("Workbook is empty.") from exc

        headers = self._make_unique_headers(raw_headers)
        missing = sorted(REQUIRED_HEADERS - set(headers))
        if missing:
            raise ValueError(f"Missing required columns: {', '.join(missing)}")

        villages_by_code = {
            village.acs_code: village.id
            for village in self.db.scalars(select(Village)).all()
        }
        conditions_by_action = self._conditions_by_action()
        sources_by_name = self._sources_by_name()

        processed = 0
        succeeded = 0
        row_errors: list[WorkbookImportRowErrorDTO] = []

        for row_number, row in enumerate(rows, start=2):
            row_data = self._row_to_record(headers, row)
            if not self._row_has_values(row_data):
                continue

            processed += 1
            try:
                incident = Incident(
                    raw_message_id=None,
                    village_id=self._resolve_village_id(row_data, villages_by_code),
                    condition_id=self._resolve_condition_id(row_data, conditions_by_action),
                    source_id=self._resolve_source_id(row_data, sources_by_name),
                    event_month=self._optional_string(row_data.get("Month")),
                    event_date=self._date(row_data.get("Date")),
                    event_time=self._time(row_data.get("Time")),
                    khabar=self._required_string(row_data.get("Khabar"), "Khabar"),
                    moh=self._optional_string(row_data.get("MOH")),
                    martyrs=self._optional_string(row_data.get("Martyrs")),
                    worker_name=self._optional_string(row_data.get("Worker Name")),
                    source_link=self._optional_string(row_data.get("Links")),
                    source_link_2=self._optional_string(row_data.get("Links__2")),
                    note=self._optional_string(row_data.get("NOTE")),
                    note_extra=self._optional_string(row_data.get("Note")),
                    note_extra_2=self._optional_string(row_data.get("Note__2")),
                    created_by=created_by,
                )
                for header, field_name in INCIDENT_INT_FIELD_MAP.items():
                    setattr(incident, field_name, self._optional_int(row_data.get(header)))

                self.db.add(incident)
                self.db.flush()

                detail_values = self._build_detail_values(row_data)
                self.db.add(IncidentDetail(incident_id=incident.id, **detail_values))
                self.db.commit()
                succeeded += 1
            except Exception as exc:
                self.db.rollback()
                row_errors.append(
                    WorkbookImportRowErrorDTO(row=row_number, error=self._format_exception(exc))
                )

        return WorkbookImportSummaryDTO(
            processed=processed,
            succeeded=succeeded,
            failed=len(row_errors),
            row_errors=row_errors,
        )

    def _ensure_schema_compatible(self) -> None:
        bind_getter = getattr(self.db, "get_bind", None)
        if bind_getter is None:
            return

        bind = bind_getter()
        if bind is None:
            return

        incident_columns = {
            column_info["name"]
            for column_info in inspect(bind).get_columns(Incident.__tablename__)
        }
        missing_columns = sorted({"note_extra_2"} - incident_columns)
        if missing_columns:
            raise ValueError(
                "Database schema is out of date for incident imports. "
                f"Missing incidents columns: {', '.join(missing_columns)}. "
                "Apply the latest Alembic migrations before importing incidents."
            )

    def _conditions_by_action(self) -> dict[str, int]:
        conditions_by_action: dict[str, int] = {}
        for condition in self.db.scalars(select(Condition)).all():
            for value in (condition.action_en, condition.action_ar):
                key = self._normalized_lookup(value)
                if key:
                    conditions_by_action[key] = condition.id
        return conditions_by_action

    def _sources_by_name(self) -> dict[str, int]:
        sources_by_name: dict[str, int] = {}
        for source in self.db.scalars(select(Source)).all():
            key = self._normalized_lookup(source.name)
            if key:
                sources_by_name[key] = source.id
        return sources_by_name

    def _resolve_village_id(self, row_data: dict[str, Any], villages_by_code: dict[int, int]) -> int | None:
        acs_code = self._optional_int(row_data.get("ACS_Code"))
        if acs_code is None:
            return None
        return villages_by_code.get(acs_code)

    def _resolve_condition_id(
        self,
        row_data: dict[str, Any],
        conditions_by_action: dict[str, int],
    ) -> int | None:
        action_en = self._normalized_lookup(row_data.get("Action_E"))
        action_ar = self._normalized_lookup(row_data.get("Action_A"))
        return conditions_by_action.get(action_en) or conditions_by_action.get(action_ar)

    def _resolve_source_id(self, row_data: dict[str, Any], sources_by_name: dict[str, int]) -> int | None:
        return sources_by_name.get(self._normalized_lookup(row_data.get("Source")))

    def _build_detail_values(self, row_data: dict[str, Any]) -> dict[str, Any]:
        values: dict[str, Any] = {}
        for header, field_name in INCIDENT_DETAIL_FIELD_MAP.items():
            column = IncidentDetail.__table__.columns[field_name]
            values[field_name] = self._coerce_detail_value(column.type, row_data.get(header), header)
        return values

    @staticmethod
    def _make_unique_headers(headers: tuple[Any, ...] | list[Any] | Any) -> list[str]:
        counts: dict[str, int] = {}
        unique_headers: list[str] = []
        for index, value in enumerate(headers, start=1):
            base = str(value).strip() if value is not None else ""
            base = base or f"Column_{index}"
            counts[base] = counts.get(base, 0) + 1
            if counts[base] == 1:
                unique_headers.append(base)
            else:
                unique_headers.append(f"{base}__{counts[base]}")
        return unique_headers

    @staticmethod
    def _row_to_record(headers: list[str], row: tuple[Any, ...]) -> dict[str, Any]:
        return {
            headers[index]: row[index] if index < len(row) else None
            for index in range(len(headers))
        }

    @staticmethod
    def _row_has_values(row_data: dict[str, Any]) -> bool:
        return any(value not in (None, "") for value in row_data.values())

    @staticmethod
    def _required_string(value: Any, field_name: str) -> str:
        text = IncidentWorkbookService._optional_string(value)
        if text is None:
            raise ValueError(f"{field_name} is required.")
        return text

    @staticmethod
    def _optional_string(value: Any) -> str | None:
        if value in (None, ""):
            return None
        text = str(value).strip()
        return text or None

    @staticmethod
    def _optional_int(value: Any) -> int | None:
        if value in (None, ""):
            return None
        if isinstance(value, bool):
            return int(value)
        if isinstance(value, int):
            return value
        if isinstance(value, float):
            if not value.is_integer():
                raise ValueError(f"Expected an integer value, got {value!r}.")
            return int(value)

        text = str(value).strip()
        if not text:
            return None
        number = float(text)
        if not number.is_integer():
            raise ValueError(f"Expected an integer value, got {text!r}.")
        return int(number)

    @staticmethod
    def _date(value: Any) -> date:
        if isinstance(value, datetime):
            return value.date()
        if isinstance(value, date):
            return value
        text = IncidentWorkbookService._optional_string(value)
        if text is None:
            raise ValueError("Date is required.")
        return datetime.fromisoformat(text).date()

    @staticmethod
    def _time(value: Any) -> time | None:
        if value in (None, ""):
            return None
        if isinstance(value, datetime):
            return value.time().replace(tzinfo=None)
        if isinstance(value, time):
            return value.replace(tzinfo=None)
        text = IncidentWorkbookService._optional_string(value)
        if text is None:
            return None
        return time.fromisoformat(text)

    @staticmethod
    def _normalized_lookup(value: Any) -> str:
        if value in (None, ""):
            return ""
        return str(value).strip().casefold()

    @staticmethod
    def _coerce_detail_value(column_type: Any, value: Any, header: str) -> Any:
        if isinstance(column_type, Boolean):
            return IncidentWorkbookService._optional_bool(value)
        if isinstance(column_type, Integer):
            return IncidentWorkbookService._optional_int(value)
        if isinstance(column_type, Enum):
            return IncidentWorkbookService._optional_did(value, header)
        return IncidentWorkbookService._optional_string(value)

    @staticmethod
    def _optional_bool(value: Any) -> bool | None:
        if value in (None, ""):
            return None
        if isinstance(value, bool):
            return value
        if isinstance(value, (int, float)):
            return bool(value)

        text = str(value).strip().casefold()
        if text in {"1", "1.0", "true", "yes", "y"}:
            return True
        if text in {"0", "0.0", "false", "no", "n"}:
            return False
        raise ValueError(f"Could not parse boolean value {value!r}.")

    @staticmethod
    def _optional_did(value: Any, header: str) -> str | None:
        text = IncidentWorkbookService._optional_string(value)
        if text is None:
            return None
        normalized = text.upper()
        if normalized not in {"D", "ID"}:
            raise ValueError(f"{header} must be 'D' or 'ID'.")
        return normalized

    @staticmethod
    def _format_exception(exc: Exception) -> str:
        message = str(exc).strip()
        if message:
            return f"{type(exc).__name__}: {message}"
        return f"{type(exc).__name__} (no message)"
